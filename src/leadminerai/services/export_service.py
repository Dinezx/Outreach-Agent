from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook

from leadminerai.schemas.company import CompanyRead


class ExportService:
    @staticmethod
    def build_excel(companies: list[CompanyRead]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "companies"
        worksheet.append(["id", "name", "website_url", "status", "last_error", "created_at", "updated_at"])

        for company in companies:
            worksheet.append(
                [
                    company.id,
                    company.name,
                    company.website_url,
                    company.status.value,
                    company.last_error,
                    company.created_at.isoformat(),
                    company.updated_at.isoformat(),
                ]
            )

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def build_contacts_excel(contacts: list[dict]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "contacts"
        headers = [
            "company_name", "email", "phone", "mobile", "address", "city", "state",
            "country", "linkedin", "facebook", "instagram", "youtube", "contact_page",
            "maps_url", "confidence_score", "created_at", "updated_at"
        ]
        worksheet.append(headers)

        for contact in contacts:
            worksheet.append([contact.get(h) for h in headers])

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def build_contacts_csv(contacts: list[dict]) -> bytes:
        import csv
        from io import StringIO
        
        output = StringIO()
        headers = [
            "company_name", "email", "phone", "mobile", "address", "city", "state",
            "country", "linkedin", "facebook", "instagram", "youtube", "contact_page",
            "maps_url", "confidence_score", "created_at", "updated_at"
        ]
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for contact in contacts:
            row = {}
            for h in headers:
                val = contact.get(h)
                if isinstance(val, datetime):
                    row[h] = val.isoformat()
                elif val is None:
                    row[h] = ""
                else:
                    row[h] = str(val)
            writer.writerow(row)
            
        return output.getvalue().encode("utf-8")

    @staticmethod
    def build_intelligence_excel(contacts: list[dict], dms: list[dict]) -> bytes:
        workbook = Workbook()
        
        # Contacts sheet
        ws_contacts = workbook.active
        ws_contacts.title = "Contacts"
        contact_headers = ["company_name", "contact_type", "contact_value", "contact_label", "priority", "confidence", "source_url", "created_at"]
        ws_contacts.append(contact_headers)
        for c in contacts:
            ws_contacts.append([
                c.get("company_name"),
                c.get("contact_type"),
                c.get("contact_value"),
                c.get("contact_label"),
                c.get("priority"),
                c.get("confidence"),
                c.get("source_url"),
                c.get("created_at").isoformat() if isinstance(c.get("created_at"), datetime) else str(c.get("created_at") or "")
            ])

        # Decision Makers sheet
        ws_dms = workbook.create_sheet(title="Decision Makers")
        dm_headers = ["company_name", "name", "designation", "linkedin_url", "priority", "confidence", "source_url", "created_at"]
        ws_dms.append(dm_headers)
        for d in dms:
            ws_dms.append([
                d.get("company_name"),
                d.get("name"),
                d.get("designation"),
                d.get("linkedin_url"),
                d.get("priority"),
                d.get("confidence"),
                d.get("source_url"),
                d.get("created_at").isoformat() if isinstance(d.get("created_at"), datetime) else str(d.get("created_at") or "")
            ])

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def build_intelligence_csv(contacts: list[dict], dms: list[dict]) -> bytes:
        import csv
        from io import StringIO
        
        output = StringIO()
        headers = ["company_name", "record_type", "name_or_value", "label_or_designation", "linkedin_url", "priority", "confidence", "source_url", "created_at"]
        writer = csv.writer(output)
        writer.writerow(headers)
        
        # Write contacts
        for c in contacts:
            created = c.get("created_at")
            created_str = created.isoformat() if isinstance(created, datetime) else str(created or "")
            writer.writerow([
                c.get("company_name"),
                "Contact",
                c.get("contact_value"),
                c.get("contact_label"),
                "",  # linkedin_url not applicable for generic contacts
                c.get("priority"),
                c.get("confidence"),
                c.get("source_url"),
                created_str
            ])
            
        # Write decision makers
        for d in dms:
            created = d.get("created_at")
            created_str = created.isoformat() if isinstance(created, datetime) else str(created or "")
            writer.writerow([
                d.get("company_name"),
                "Decision Maker",
                d.get("name"),
                d.get("designation"),
                d.get("linkedin_url"),
                d.get("priority"),
                d.get("confidence"),
                d.get("source_url"),
                created_str
            ])
            
        return output.getvalue().encode("utf-8")

    @staticmethod
    def build_business_intelligence_excel(items: list[dict]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Business Intelligence"
        headers = [
            "company_name", "industry", "sub_industry", "manufacturing_type",
            "products", "services", "certifications", "locations", "markets",
            "predicted_departments", "predicted_pain_points", "confidence", "description"
        ]
        worksheet.append(headers)

        for item in items:
            products_str = ", ".join(item.get("products") or [])
            services_str = ", ".join(item.get("services") or [])
            certs_str = ", ".join(item.get("certifications") or [])
            locs_str = ", ".join(item.get("locations") or [])
            markets_str = ", ".join(item.get("markets") or [])
            depts_str = ", ".join([
                f"{d['name']} ({d['confidence']}%)" if isinstance(d, dict) else str(d)
                for d in (item.get("departments") or [])
            ])
            pains_str = " | ".join([
                f"{p['name']} [Sev:{p.get('severity',0)}, Freq:{p.get('frequency','')}]" if isinstance(p, dict) else str(p)
                for p in (item.get("predicted_pain_points") or [])
            ])

            worksheet.append([
                item.get("company_name"),
                item.get("industry"),
                item.get("sub_industry"),
                item.get("manufacturing_type"),
                products_str,
                services_str,
                certs_str,
                locs_str,
                markets_str,
                depts_str,
                pains_str,
                item.get("confidence"),
                item.get("description")
            ])

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def build_business_intelligence_csv(items: list[dict]) -> bytes:
        import csv
        from io import StringIO

        output = StringIO()
        headers = [
            "company_name", "industry", "sub_industry", "manufacturing_type",
            "products", "services", "certifications", "locations", "markets",
            "predicted_departments", "predicted_pain_points", "confidence", "description"
        ]
        writer = csv.writer(output)
        writer.writerow(headers)

        for item in items:
            products_str = ", ".join(item.get("products") or [])
            services_str = ", ".join(item.get("services") or [])
            certs_str = ", ".join(item.get("certifications") or [])
            locs_str = ", ".join(item.get("locations") or [])
            markets_str = ", ".join(item.get("markets") or [])
            depts_str = ", ".join([
                f"{d['name']} ({d['confidence']}%)" if isinstance(d, dict) else str(d)
                for d in (item.get("departments") or [])
            ])
            pains_str = " | ".join([
                f"{p['name']} [Sev:{p.get('severity',0)}, Freq:{p.get('frequency','')}]" if isinstance(p, dict) else str(p)
                for p in (item.get("predicted_pain_points") or [])
            ])

            writer.writerow([
                item.get("company_name"),
                item.get("industry"),
                item.get("sub_industry"),
                item.get("manufacturing_type"),
                products_str,
                services_str,
                certs_str,
                locs_str,
                markets_str,
                depts_str,
                pains_str,
                item.get("confidence"),
                item.get("description")
            ])

        return output.getvalue().encode("utf-8")

    @staticmethod
    def build_business_intelligence_json(items: list[dict]) -> bytes:
        import json
        return json.dumps(items, indent=2, default=str).encode("utf-8")

    @staticmethod
    def build_outreach_excel(campaigns: list[dict]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Outreach Campaigns"
        headers = [
            "id", "company_name", "target_role", "decision_maker_name", "channel",
            "status", "subject", "email_body", "linkedin_message", "phone_script",
            "channel_confidence", "overall_confidence", "recommendation_reason",
            "scheduled_at", "sent_at", "created_at"
        ]
        worksheet.append(headers)

        for c in campaigns:
            worksheet.append([
                c.get("id"),
                c.get("company_name"),
                c.get("target_role"),
                c.get("decision_maker_name"),
                c.get("channel"),
                str(c.get("status")),
                c.get("subject"),
                c.get("email_body"),
                c.get("linkedin_message"),
                c.get("phone_script"),
                c.get("channel_confidence"),
                c.get("overall_confidence"),
                c.get("recommendation_reason"),
                str(c.get("scheduled_at") or ""),
                str(c.get("sent_at") or ""),
                str(c.get("created_at") or "")
            ])

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def build_outreach_csv(campaigns: list[dict]) -> bytes:
        import csv
        from io import StringIO

        output = StringIO()
        headers = [
            "id", "company_name", "target_role", "decision_maker_name", "channel",
            "status", "subject", "email_body", "linkedin_message", "phone_script",
            "channel_confidence", "overall_confidence", "recommendation_reason",
            "scheduled_at", "sent_at", "created_at"
        ]
        writer = csv.writer(output)
        writer.writerow(headers)

        for c in campaigns:
            writer.writerow([
                c.get("id"),
                c.get("company_name"),
                c.get("target_role"),
                c.get("decision_maker_name"),
                c.get("channel"),
                str(c.get("status")),
                c.get("subject"),
                c.get("email_body"),
                c.get("linkedin_message"),
                c.get("phone_script"),
                c.get("channel_confidence"),
                c.get("overall_confidence"),
                c.get("recommendation_reason"),
                str(c.get("scheduled_at") or ""),
                str(c.get("sent_at") or ""),
                str(c.get("created_at") or "")
            ])

        return output.getvalue().encode("utf-8")

    @staticmethod
    def build_outreach_pdf(campaigns: list[dict]) -> bytes:
        """
        Generates a clean text/HTML formatted PDF report buffer for outreach campaigns.
        """
        from io import BytesIO

        html_content = """<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<title>LeadMiner AI - Outreach Intelligence Brief</title>
<style>
body { font-family: Helvetica, Arial, sans-serif; margin: 30px; color: #1a202c; font-size: 13px; line-height: 1.5; }
h1 { color: #0f172a; border-bottom: 2px solid #3b82f6; padding-bottom: 8px; font-size: 22px; }
.campaign-card { border: 1px solid #cbd5e1; border-radius: 8px; margin-bottom: 24px; padding: 16px; page-break-inside: avoid; }
.header-row { display: flex; justify-content: space-between; border-bottom: 1px solid #e2e8f0; padding-bottom: 8px; margin-bottom: 12px; }
.company-name { font-weight: bold; font-size: 16px; color: #1e3a8a; }
.badge { background: #eff6ff; color: #1d4ed8; padding: 3px 8px; border-radius: 4px; font-weight: bold; font-size: 11px; }
.section-title { font-weight: bold; color: #475569; text-transform: uppercase; font-size: 11px; margin-top: 12px; margin-bottom: 4px; border-bottom: 1px dashed #cbd5e1; }
.body-box { background: #f8fafc; border: 1px solid #e2e8f0; padding: 10px; border-radius: 6px; white-space: pre-wrap; font-family: monospace; font-size: 12px; }
</style>
</head>
<body>
<h1>LeadMiner AI - Industrial Research Outreach Brief</h1>
"""
        for c in campaigns:
            html_content += f"""
<div class="campaign-card">
  <div class="header-row">
    <div>
      <span class="company-name">{c.get('company_name', 'Company')}</span> — <span style="color:#64748b;">Target Role: {c.get('target_role', 'Operations')}</span>
    </div>
    <div>
      <span class="badge">Status: {c.get('status')}</span>
      <span class="badge" style="background:#f0fdf4; color:#15803d;">Confidence: {c.get('overall_confidence')}%</span>
    </div>
  </div>

  <div><strong>Recommended Channel:</strong> {c.get('channel')} (Confidence: {c.get('channel_confidence')}%)</div>
  <div style="font-style:italic; color:#64748b; font-size:12px; margin-top:4px;"><strong>Reasoning:</strong> {c.get('recommendation_reason')}</div>

  <div class="section-title">1. Research Email Draft</div>
  <div><strong>Subject:</strong> {c.get('subject')}</div>
  <div class="body-box">{c.get('email_body')}</div>

  <div class="section-title">2. LinkedIn Research Invitation (Max 300 Chars)</div>
  <div class="body-box">{c.get('linkedin_message')}</div>

  <div class="section-title">3. Phone Script (Research Call)</div>
  <div class="body-box">{c.get('phone_script')}</div>
</div>
"""
        html_content += "</body></html>"
        return html_content.encode("utf-8")



