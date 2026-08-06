from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook


def parse_company_csv(content: bytes) -> list[str]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if reader.fieldnames is None:
        raise ValueError("CSV file is missing a header row")

    header_map = {field.lower().strip(): field for field in reader.fieldnames}
    column_name = header_map.get("company_name") or header_map.get("name")
    if column_name is None:
        raise ValueError("CSV must contain a 'company_name' or 'name' column")

    names: list[str] = []
    for row in reader:
        value = (row.get(column_name) or "").strip()
        if value:
            names.append(value)
    return names


def parse_company_xlsx(content: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]

    if not rows:
        raise ValueError("XLSX file is empty")

    header_row = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    header_map = {field: index for index, field in enumerate(header_row) if field}
    column_index = header_map.get("company_name") if "company_name" in header_map else header_map.get("name")
    if column_index is None:
        raise ValueError("XLSX must contain a 'company_name' or 'name' column")

    names: list[str] = []
    for row in rows[1:]:
        if column_index >= len(row):
            continue
        value = row[column_index]
        if value is None:
            continue
        cleaned = str(value).strip()
        if cleaned:
            names.append(cleaned)
    return names


def parse_company_upload(filename: str, content: bytes) -> list[str]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return parse_company_csv(content)
    if lowered.endswith(".xlsx"):
        return parse_company_xlsx(content)
    raise ValueError("Only CSV or XLSX uploads are supported")
